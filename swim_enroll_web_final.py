
import io
from zipfile import ZipFile, ZIP_DEFLATED
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation as _DV

st.set_page_config(page_title="游泳課報名（最終版）", page_icon="🏊", layout="wide")
st.title("🏊 游泳課報名（最終版）")
st.caption("依目前規格：上傳名單 → 線上設定『參加/不參加』與（參加時）『級數 0–5』 → 一鍵匯出。匯出欄位：班級、座號、姓名、參加意願(1/0)、0級、1級、2級、3級、4級、5級；最後一列為合計。")

REQUIRED_BASE = ["班級", "座號", "姓名"]
JOIN_OPTIONS = ["參加", "不參加"]
LEVELS = [str(i) for i in range(0,6)]

INFO_MD = """
**上傳欄位**：至少含「班級、座號、姓名」。若有「參加意願」「級數」也會讀入（自動修正格式）。  
**規則**：  
- 參加意願：預設「參加」，匯出轉為 **1/0**（1=參加，0=不參加）。  
- 只有選「參加」時需填級數（0–5）。  
- 匯出檔每班一份，欄位固定：**班級、座號、姓名、參加意願、0級、1級、2級、3級、4級、5級**。  
- 最後一列為 **合計**：參加人數與各級（0~5）人數。  
"""

with st.expander("📄 使用說明 / 規格", expanded=False):
    st.markdown(INFO_MD)

# ===== 上傳檔案 =====
uploaded = st.file_uploader("上傳 Excel（.xlsx）", type=["xlsx"])

# 提供樣板下載（含下拉）
with st.expander("需要 Excel 樣板？點此下載", expanded=False):
    import pandas as _pd
    tmp = _pd.DataFrame({"班級":["四年一班"], "座號":[1], "姓名":["王小明"], "參加意願":["參加"], "級數":[""]})
    bio = io.BytesIO()
    with _pd.ExcelWriter(bio, engine="openpyxl") as writer:
        tmp.to_excel(writer, index=False, sheet_name="名單")
    bio.seek(0)
    wb = load_workbook(bio)
    ws = wb.active
    header_map = {cell.value: cell.column for cell in ws[1] if cell.value}
    col_join = header_map.get("參加意願")
    col_lv = header_map.get("級數")
    if col_join:
        dvj = _DV(type="list", formula1='"參加,不參加"', allow_blank=True)
        ws.add_data_validation(dvj)
        dvj.add(f"{ws.cell(row=2, column=col_join).coordinate}:{ws.cell(row=501, column=col_join).coordinate}")
    if col_lv:
        dvl = _DV(type="list", formula1='"0,1,2,3,4,5"', allow_blank=True)
        ws.add_data_validation(dvl)
        dvl.add(f"{ws.cell(row=2, column=col_lv).coordinate}:{ws.cell(row=501, column=col_lv).coordinate}")
    out = io.BytesIO()
    wb.save(out)
    st.download_button("⬇️ 下載樣板.xlsx", data=out.getvalue(),
        file_name="學生名單_樣板.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if not uploaded:
    st.info("請先上傳名單檔案（或下載樣板填寫後再上傳）。")
    st.stop()

# ===== 讀檔 =====
try:
    df = pd.read_excel(uploaded)
except Exception as e:
    st.error(f"讀取 Excel 失敗：{e}")
    st.stop()

df = df.fillna("")

# 欄位矯正
if "是否參加" in df.columns and "參加意願" not in df.columns:
    df = df.rename(columns={"是否參加":"參加意願"})
for col in REQUIRED_BASE:
    if col not in df.columns:
        df[col] = ""
if "參加意願" not in df.columns:
    df["參加意願"] = ""
if "級數" not in df.columns:
    df["級數"] = ""

# 值規範
df["參加意願"] = df["參加意願"].replace({"是":"參加","否":"不參加"})
df.loc[df["參加意願"].astype(str).str.strip() == "", "參加意願"] = "參加"
df.loc[~df["級數"].astype(str).isin(LEVELS), "級數"] = ""

# ===== 篩選與編輯 =====
classes = sorted([c for c in df["班級"].astype(str).unique() if c.strip()])
left, right = st.columns([1,2])
sel_class = left.selectbox("選擇班級", options=classes if classes else [""])
name_filter = right.text_input("姓名關鍵字（可留白）")

work_df = df.copy()

st.subheader("名單編輯")
st.write("提示：**參加**時才需選擇級數（0–5）。可逐列設定。")

# 表頭
header_cols = st.columns([1,2,1.2,1.2])
header_cols[0].markdown("**座號**")
header_cols[1].markdown("**姓名**")
header_cols[2].markdown("**參加意願**")
header_cols[3].markdown("**級數**")

mask = (work_df["班級"].astype(str) == sel_class)
if name_filter.strip():
    mask &= work_df["姓名"].astype(str).str.contains(name_filter.strip(), case=False, na=False)
view_df = work_df[mask]

edited_rows = []
for idx, row in view_df.iterrows():
    c = st.columns([1,2,1.2,1.2])
    c[0].text(str(row.get("座號","")))
    c[1].text(str(row.get("姓名","")))

    join_val = str(row.get("參加意願","參加"))
    if join_val not in JOIN_OPTIONS:
        join_val = "參加"
    join_sel = c[2].selectbox(f"參加意願_{idx}", JOIN_OPTIONS, index=JOIN_OPTIONS.index(join_val), key=f"join_{idx}")
    if join_sel == "參加":
        level_val = str(row.get("級數",""))
        if level_val not in LEVELS:
            level_val = ""
        level_sel = c[3].selectbox(f"級數_{idx}", [""] + LEVELS, index=([""]+LEVELS).index(level_val) if level_val in ([""]+LEVELS) else 0, key=f"lv_{idx}")
    else:
        level_sel = ""
        c[3].text("")

    edited = row.copy()
    edited["參加意願"] = join_sel
    edited["級數"] = level_sel
    edited_rows.append((idx, edited))

# 回寫
for idx, row in edited_rows:
    work_df.loc[idx, "參加意願"] = row["參加意願"]
    work_df.loc[idx, "級數"] = row["級數"]

# ===== 統計 =====
st.divider()
st.subheader("即時統計")
sub = work_df[work_df["班級"].astype(str) == sel_class]
total = len(sub)
joined = (sub["參加意願"]=="參加").sum()
not_joined = (sub["參加意願"]=="不參加").sum()
level_counts = (
    sub.loc[(sub["參加意願"]=="參加") & (sub["級數"].isin(LEVELS)), "級數"]
    .value_counts().reindex(LEVELS, fill_value=0)
)
c1,c2,c3,c4,c5,c6,c7 = st.columns(7)
c1.metric("總人數", total)
c2.metric("參加", joined)
c3.metric("不參加", not_joined)
c4.metric("0級", int(level_counts.get("0",0)))
c5.metric("1級", int(level_counts.get("1",0)))
c6.metric("2級", int(level_counts.get("2",0)))
c7.metric("3-5級", int(level_counts.get("3",0)+level_counts.get("4",0)+level_counts.get("5",0)))
st.bar_chart(level_counts)

# ===== 匯出 =====
st.divider()
st.subheader("匯出")

def make_class_excel(df_class: pd.DataFrame) -> bytes:
    out = df_class.copy()
    out["參加意願"] = out["參加意願"].map({"參加":1,"不參加":0}).fillna(1).astype(int)

    # 0~5 級 one-hot
    out["0級"] = ((df_class["級數"].astype(str) == "0") & (out["參加意願"]==1)).astype(int)
    for i in range(1,6):
        col = f"{i}級"
        out[col] = ((df_class["級數"].astype(str) == str(i)) & (out["參加意願"]==1)).astype(int)

    # 欄位順序
    out = out[["班級","座號","姓名","參加意願","0級","1級","2級","3級","4級","5級"]]

    # 排序（座號數字優先）
    def to_num(x):
        try:
            return float(x)
        except:
            return float('inf')
    out = out.sort_values(by=["班級","座號"], key=lambda col: col.map(to_num) if col.name=="座號" else col)

    # 合計列
    total_row = {col:"" for col in out.columns}
    total_row["班級"] = "合計"
    total_row["參加意願"] = int(out["參加意願"].sum())
    for col in ["0級","1級","2級","3級","4級","5級"]:
        total_row[col] = int(out[col].sum())
    out = pd.concat([out, pd.DataFrame([total_row])], ignore_index=True)

    # 輸出
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        out.to_excel(writer, index=False, sheet_name=str(df_class['班級'].iloc[0]))
    return bio.getvalue()

# 單班下載
if classes and sel_class:
    df_class = work_df[work_df["班級"].astype(str) == sel_class][["班級","座號","姓名","參加意願","級數"]].copy()
    st.download_button(
        label=f"⬇️ 下載 {sel_class} 名單（Excel）",
        data=make_class_excel(df_class),
        file_name=f"{sel_class}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# 全部班級 ZIP（直接提供按鈕，不另外用 st.button 以避免狀態混淆）
if classes:
    zip_io = io.BytesIO()
    with ZipFile(zip_io, "w", ZIP_DEFLATED) as zf:
        for c in classes:
            subc = work_df[work_df["班級"].astype(str) == c][["班級","座號","姓名","參加意願","級數"]]
            if len(subc)==0:
                continue
            content = make_class_excel(subc)
            zf.writestr(f"{c}.xlsx", content)
    st.download_button(
        label="⬇️ 下載全部班級（ZIP）",
        data=zip_io.getvalue(),
        file_name="全部班級_游泳課名單.zip",
        mime="application/zip"
    )

st.success("完成：符合最新欄位與合計規格，支援每班Excel與全班級ZIP匯出。")
